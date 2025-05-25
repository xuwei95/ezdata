from web_apps.rag.extractor.extractor_base import BaseExtractor
from langchain_core.documents import Document
import logging
from typing import Optional
import pandas as pd
from openpyxl.reader.excel import load_workbook
from pandas import DataFrame
from web_apps.rag.extractor.table_utils import split_to_multi_document_tables, extract_tables_by_chunk_max_token

_punctuation = r"""!"#$%&'()*+,-./:;<=>?@[\]^_`{|}~！“”‘’（），。：；、？》《」「【】"""
sub_theme_name = 'sub_theme_name'


def _convert_to_html(table: list) -> str:
    # Helper function to convert a 2D list representing a table into an HTML table string.
    html = "<table>"
    for row in table:
        html += "<tr>"
        for cell in row:
            html += f"<td>{cell}</td>"
        html += "</tr>"
    html += "</table>"
    return html


def _handler_merged_cell_df(sheet, df: DataFrame, header=None) -> DataFrame:
    '''
    处理合并单元格。将合并单元格内的值设置为合并单元格的值（即start_cell的值）
    :param sheet:
    :param df:
    :param header:
    :return:
    '''
    for item in sheet.merged_cells:
        top_col, top_row, bottom_col, bottom_row = item.bounds
        base_value = item.start_cell.value
        # 1-based index转为0-based index
        top_row -= 1
        top_col -= 1
        # 由于前面的几行被设为了header，所以这里要对坐标进行调整(因为df的二维数组从0算，而sheet的cells还是按整个表格算的)
        if header is not None:
            top_row -= header + 1
            bottom_row -= header + 1
        if top_row < 0 or bottom_row < 0:
            # 如果减成负数了，说明是属于header以及header上面有合并单元格，这部分我们只取了header，上面的数据我们不管
            continue
        df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value
    return df


def _find_df_real_header(df: DataFrame, start_header=0, max_header=2):
    tmp_header = start_header
    headers = [hd for hd in df.columns.tolist() if "Unnamed:" not in hd]
    if len(headers) == len(df.columns) or start_header >= max_header:
        return tmp_header
    else:
        tmp_header = tmp_header + 1
        for row in df.values.tolist():
            tmp_row = [r for r in row if not pd.isna(r)]
            if len(tmp_row) < len(df.columns) and tmp_header < max_header:
                tmp_header = tmp_header + 1
            else:
                break
        return tmp_header


def _parse_df_header_and_values(df: DataFrame):
    headers = df.columns.tolist()
    values = df.values.tolist()
    datas = []
    for row_idx, row in enumerate(values):
        row_data = []
        for col_idx, col in enumerate(row):
            col_value = str(col)
            if col_value == 'nan':
                col_value = ''
            row_data.append(col_value.strip())
        datas.append(row_data)
    return headers, datas


def _find_sheet_real_max_rows_columns(sheet):
    # 允许每个sheet最大的行数和列数
    sheet_max_row = min(sheet.max_row, int("100000"))
    sheet_max_column = min(sheet.max_column, int("100"))

    # 允许最长的连续空白的行和列数，超过这个的行和列不再读取
    tolerance_continue_row_blank = int("20")
    tolerance_continue_column_blank = int("10")

    def _invalid_value(v):
        return (v is None
                or str(v).strip() == ''
                or str(v).strip() == 'nan'
                or (len(str(v).strip()) == 1 and str(v).strip() in _punctuation))

    max_row = 0
    max_columns = []
    continue_row_blank = 0
    for i in range(1, sheet_max_row + 1):

        continue_col_blank = 0
        row_max_column = 0
        for j in range(1, sheet_max_column + 1):
            cell = sheet.cell(row=i, column=j)
            if _invalid_value(cell.value):
                continue_col_blank = continue_col_blank + 1
            else:
                row_max_column = j
                continue_col_blank = 0

            if continue_col_blank >= tolerance_continue_column_blank:
                break
        max_columns.append(row_max_column)

        if row_max_column == 0:
            continue_row_blank = continue_row_blank + 1
        else:
            max_row = i
            continue_row_blank = 0

        if continue_row_blank >= tolerance_continue_row_blank:
            break

    max_column = max(max_columns)
    logging.info(f"sheet({sheet.title}):[{sheet.max_row},{sheet.max_column},real:[{max_row},{max_column}]")
    return max_row, max_column


def read_xlsx(file):
    excel = pd.ExcelFile(load_workbook(file), engine="openpyxl")
    for sheet_name in excel.sheet_names:
        sheet = excel.book[sheet_name]
        find_max_row, find_max_column = _find_sheet_real_max_rows_columns(sheet)
        header = 0
        df = excel.parse(sheet_name, header=header, nrows=find_max_row, usecols=range(0, find_max_column))
        max_header = int("6")
        header = _find_df_real_header(df, start_header=header, max_header=max_header)
        if header > 0:
            df = excel.parse(sheet_name, header=header, nrows=find_max_row, usecols=range(0, find_max_column))
        df = _handler_merged_cell_df(sheet, df, header=header)
        headers, values = _parse_df_header_and_values(df)
        yield sheet_name, headers, values


class ExcelExtractor(BaseExtractor):
    """Load Excel files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(
            self,
            file_path: str,
            encoding: Optional[str] = None,
            autodetect_encoding: bool = False
    ):
        """Initialize with file path."""
        self._file_path = file_path
        self._encoding = encoding
        self._autodetect_encoding = autodetect_encoding

    def extract(self) -> list[Document]:
        """ Load from Excel file in xls or xlsx format using Pandas."""
        documents = []
        for sheet_name, headers, values in read_xlsx(self._file_path):
            # 空列去除(以header为主对齐)
            useful_columns = 0
            for idx, header in enumerate(headers):
                if str(header).startswith('Unnamed:'):
                    useful_columns = idx
                    break
            if not useful_columns:
                useful_columns = len(headers)
            headers = headers[:useful_columns]
            values = [data[:useful_columns] for data in values]

            array_2d_list_with_headers = [headers] + values
            table_htmls = extract_tables_by_chunk_max_token(table_array=array_2d_list_with_headers,
                                                            table_max_token=1024) + '\n'
            sheet_documents = split_to_multi_document_tables(text=table_htmls, metadata={sub_theme_name: sheet_name})
            documents.extend(sheet_documents)
        return documents


if __name__ == '__main__':
    extractor = ExcelExtractor('test.xlsx')
    documents = extractor.extract()
    print(documents)


