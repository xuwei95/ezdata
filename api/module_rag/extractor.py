"""
文档抽取 —— 按扩展名把文件解析成纯文本。

走 storage(utils.storage_utils)取文件,本地落临时文件再解析。
内置:txt/md/markdown/csv/tsv/json/jsonl/log(文本类)、pdf(pypdf)、xlsx/xls(openpyxl)、
      html/htm(标签清理)、docx(python-docx,装了才支持)。
缺依赖时抛出明确提示。网页/Notion 等来源在 service 层处理(此处只管文件)。
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import tempfile

_TEXT_EXTS = {'.txt', '.md', '.markdown', '.log', '.text'}


def supported_exts() -> list[str]:
    return sorted(_TEXT_EXTS | {'.csv', '.tsv', '.json', '.jsonl', '.pdf', '.xlsx', '.xls', '.html', '.htm', '.docx'})


def extract_file(file_key: str, *, filename: str | None = None) -> str:
    """从 storage 取 file_key 指向的文件并解析为文本。"""
    name = filename or os.path.basename(file_key)
    ext = os.path.splitext(name)[1].lower()
    raw = _read_storage_bytes(file_key)
    return extract_bytes(raw, ext)


def extract_bytes(raw: bytes, ext: str) -> str:
    ext = ext.lower()
    if ext in _TEXT_EXTS:
        return _decode(raw)
    if ext in ('.csv', '.tsv'):
        return _csv(raw, delimiter='\t' if ext == '.tsv' else ',')
    if ext in ('.json', '.jsonl'):
        return _json(raw, lines=ext == '.jsonl')
    if ext == '.pdf':
        return _pdf(raw)
    if ext in ('.xlsx', '.xls'):
        return _xlsx(raw)
    if ext in ('.html', '.htm'):
        return _html(raw)
    if ext == '.docx':
        return _docx(raw)
    # 兜底:按文本解码
    return _decode(raw)


# ---------------- storage ----------------
def _normalize_key(file_key: str) -> str:
    """剥掉上传接口返回的 UPLOAD_PREFIX(如 /profile/),还原为 storage 对象键。"""
    from config.env import UploadConfig  # noqa: PLC0415
    prefix = (UploadConfig.UPLOAD_PREFIX or '').strip('/')
    k = file_key.lstrip('/')
    if prefix and k.startswith(prefix + '/'):
        k = k[len(prefix) + 1:]
    return k


def _read_storage_bytes(file_key: str) -> bytes:
    from utils.storage_utils import storage  # noqa: PLC0415
    file_key = _normalize_key(file_key)
    data = storage.load_once(file_key) if hasattr(storage, 'load_once') else None
    if data is not None:
        return data if isinstance(data, bytes) else bytes(data)
    # 退化:下载到临时文件再读
    with tempfile.NamedTemporaryFile(delete=False) as tf:
        tmp = tf.name
    try:
        storage.download(file_key, tmp)
        with open(tmp, 'rb') as f:
            return f.read()
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


# ---------------- 各类型解析 ----------------
def _decode(raw: bytes) -> str:
    for enc in ('utf-8', 'utf-8-sig', 'gbk', 'gb18030', 'latin-1'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='ignore')


def _csv(raw: bytes, delimiter: str = ',') -> str:
    text = _decode(raw)
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return ''
    header = rows[0]
    lines = []
    for r in rows[1:]:
        pairs = [f'{header[i] if i < len(header) else f"col{i}"}: {v}' for i, v in enumerate(r)]
        lines.append('; '.join(pairs))
    return '\n'.join(lines) if lines else text


def _json(raw: bytes, lines: bool = False) -> str:
    text = _decode(raw)
    try:
        if lines:
            objs = [json.loads(ln) for ln in text.splitlines() if ln.strip()]
        else:
            obj = json.loads(text)
            objs = obj if isinstance(obj, list) else [obj]
        return '\n'.join(json.dumps(o, ensure_ascii=False) for o in objs)
    except Exception:
        return text


def _pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader  # noqa: PLC0415
    except ImportError as e:
        raise RuntimeError('解析 PDF 需要 pypdf,请 pip install pypdf') from e
    reader = PdfReader(io.BytesIO(raw))
    return '\n'.join((page.extract_text() or '') for page in reader.pages).strip()


def _xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as e:
        raise RuntimeError('解析 Excel 需要 openpyxl,请 pip install openpyxl') from e
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f'# {ws.title}')
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        header = [str(h) if h is not None else f'col{i}' for i, h in enumerate(header or [])]
        for r in rows:
            pairs = [f'{header[i] if i < len(header) else f"col{i}"}: {v}'
                     for i, v in enumerate(r) if v is not None]
            if pairs:
                out.append('; '.join(pairs))
    return '\n'.join(out)


def _html(raw: bytes) -> str:
    text = _decode(raw)
    try:
        from bs4 import BeautifulSoup  # noqa: PLC0415
        soup = BeautifulSoup(text, 'html.parser')
        for tag in soup(['script', 'style', 'noscript']):
            tag.decompose()
        return re.sub(r'\n{3,}', '\n\n', soup.get_text('\n')).strip()
    except ImportError:
        text = re.sub(r'(?is)<(script|style).*?>.*?</\1>', '', text)
        text = re.sub(r'(?s)<[^>]+>', ' ', text)
        return re.sub(r'[ \t]{2,}', ' ', text).strip()


def _docx(raw: bytes) -> str:
    try:
        import docx  # noqa: PLC0415
    except ImportError as e:
        raise RuntimeError('解析 docx 需要 python-docx,请 pip install python-docx') from e
    doc = docx.Document(io.BytesIO(raw))
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
