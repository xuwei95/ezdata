"""
批量导入 QA对 / 分段(CSV/Excel)→ 解析 → 批量向量化 → 入库。

CSV/Excel 表头(大小写不敏感,中英文皆可):
  QA   : question/问题, answer/答案
  分段 : content/内容(无表头时取第一列)
同步执行(自带 sync 会话 + 租户),复用 embed 缓存与向量库抽象。
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime
from typing import Any

from sqlalchemy import select

from common.context import RequestContext
from module_rag.entity.do.rag_do import RagChunk, RagDataset
from module_rag.extractor import _normalize_key, _read_storage_bytes
from module_rag.runtime_util import build_embedding_client, build_store, embed_with_cache, md5
from module_task_schedule.sync_db import get_sync_session_local

_Q_KEYS = ('question', '问题', 'q', 'query')
_A_KEYS = ('answer', '答案', 'a', 'reply')
_C_KEYS = ('content', '内容', 'text', '文本', 'chunk')


def _pick(row: dict, keys: tuple) -> str:
    low = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
    for k in keys:
        if k in low and low[k] not in (None, ''):
            return str(low[k]).strip()
    return ''


def _parse_rows(file_key: str, chunk_type: str) -> list[dict]:
    """读 storage 文件,按类型解析成 [{question,answer}] 或 [{content}]。"""
    import os

    raw = _read_storage_bytes(file_key)
    ext = os.path.splitext(_normalize_key(file_key))[1].lower()
    records: list[dict] = []

    if ext in ('.xlsx', '.xls'):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip().lower() if h is not None else f'col{i}' for i, h in enumerate(next(it, []) or [])]
        dict_rows = [dict(zip(header, r)) for r in it]
    else:  # csv / tsv / 其它按文本
        text = raw.decode('utf-8-sig', errors='ignore')
        delim = '\t' if ext == '.tsv' else ','
        dict_rows = list(csv.DictReader(io.StringIO(text), delimiter=delim))

    for row in dict_rows:
        if not row:
            continue
        if chunk_type == 'qa':
            q, a = _pick(row, _Q_KEYS), _pick(row, _A_KEYS)
            if q and a:
                records.append({'question': q, 'answer': a})
        else:
            c = _pick(row, _C_KEYS) or next((str(v).strip() for v in row.values() if v not in (None, '')), '')
            if c:
                records.append({'content': c})
    return records


def bulk_import_from_file(dataset_id: str, chunk_type: str, file_key: str, tenant_id: Any) -> dict:
    chunk_type = 'qa' if chunk_type == 'qa' else 'chunk'
    rows = _parse_rows(file_key, chunk_type)
    if not rows:
        return {'imported': 0, 'note': '未解析到有效行(检查表头:QA 需 question/answer 列,分段需 content 列)'}

    db = get_sync_session_local()()
    token = RequestContext.set_current_tenant_id(tenant_id)
    try:
        dataset = db.execute(select(RagDataset).where(RagDataset.id == dataset_id)).scalars().first()
        if dataset is None:
            return {'imported': 0, 'error': '知识库不存在'}

        texts = [r['question'] if chunk_type == 'qa' else r['content'] for r in rows]
        client = build_embedding_client(dataset)
        vectors = embed_with_cache(db, texts, dataset, client)
        dims = len(vectors[0])
        if not dataset.embedding_dims:
            from config.env import RagConfig

            dataset.embedding_dims = dims
            if not dataset.embedding_provider:
                dataset.embedding_provider = RagConfig.embedding_type
                dataset.embedding_model = RagConfig.embedding_model
            db.commit()

        store = build_store(dataset)
        store.ensure_index(dims)
        tenant_str = str(tenant_id) if tenant_id is not None else ''
        es_docs, orm_rows = [], []
        for r, vec, text in zip(rows, vectors, texts):
            cid = uuid.uuid4().hex
            es_docs.append(
                {
                    'chunk_id': cid,
                    'content': text,
                    'content_vector': vec,
                    'tenant_id': tenant_str,
                    'dataset_id': dataset.id,
                    'document_id': 'bulk',
                    'chunk_type': chunk_type,
                    **({'question': r['question']} if chunk_type == 'qa' else {}),
                }
            )
            orm_rows.append(
                RagChunk(
                    id=cid,
                    dataset_id=dataset.id,
                    document_id='bulk',
                    chunk_type=chunk_type,
                    content=(None if chunk_type == 'qa' else text),
                    question=(r['question'] if chunk_type == 'qa' else None),
                    answer=(r['answer'] if chunk_type == 'qa' else None),
                    question_hash=(md5(r['question']) if chunk_type == 'qa' else None),
                    hash=md5(text),
                    position=0,
                    status=1,
                    create_time=datetime.now(),
                )
            )
        store.add(es_docs)
        for o in orm_rows:
            db.add(o)
        db.commit()
        return {'imported': len(orm_rows)}
    except Exception as e:
        db.rollback()
        return {'imported': 0, 'error': ' '.join(str(e).split())[:500]}
    finally:
        RequestContext.reset_current_tenant_id(token)
        db.close()
